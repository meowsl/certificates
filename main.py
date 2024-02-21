import io, os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from PIL import ImageFont
from PyPDF2 import PdfReader, PdfWriter
import time
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from progress.bar import IncrementalBar, Bar

def create_cert(lastname, firstname, midname, id_text, path):

    old_pdf = PdfReader('inputs/cert.pdf')
    old_pdf_page = old_pdf.pages[0]
    size = old_pdf_page.mediabox

    packet = io.BytesIO()
    name = f'{lastname} {firstname} {midname}'
    name_width, name_height = font.getmask(name).size
    can = canvas.Canvas(packet, pagesize=letter)

    text_x = (size.width - name_width) / 2
    text_y = (size.height - name_height) / 2

    can.setFont("OpenSansItalic", 12)
    can.drawString(200, int(text_y) - 210, id_text)

    can.setFont("OpenSans", 18)
    can.drawString(int(text_x), int(text_y) - 25, name)

    packet.seek(0)
    can.save()

    new_pdf = PdfReader(packet)
    output = PdfWriter()
    page = old_pdf.pages[0]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)

    with open(f"{path}/{lastname}_{firstname[0]}_{midname[0]}.pdf", "wb") as output_stream:
        output.write(output_stream)

    bar.next()

if __name__ == "__main__":

    pdfmetrics.registerFont(TTFont('OpenSans', 'fonts/OpenSans-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('OpenSansItalic', 'fonts/OpenSans_SemiCondensed-Italic.ttf'))

    font = ImageFont.truetype("fonts/OpenSans-Regular.ttf", 18)

    start = time.time()
    wb = openpyxl.load_workbook('inputs/students.xlsx')

    for item in wb.sheetnames:
        path = f'outputs/{item}'
        if not os.path.exists(path):
            os.makedirs(path)

        sheet = wb[item]
        bar = Bar('Progress: ', max=sheet.max_row, fill='*', suffix='%(percent)d%%')
        bar.update()
        for row in sheet.iter_rows():
            elements = [str(cell.value) for cell in row]
            create_cert(elements[1], elements[2], elements[3], elements[0], path)

    print(f'Выполнилось за {time.time() - start} секунд')

